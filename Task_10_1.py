import openpyxl
from openpyxl import Workbook
#wb = Workbook()
#wb.save('Task_10-1.xlsx')
wb = openpyxl.load_workbook('Task_10-1.xlsx')
ws = wb['Sheet']
ws2 = wb['Итоги']
sotrsum = {}
for r in range(2, ws.max_row):
    sotr = ws.cell(row=r, column=1).value
    sotrsum[sotr] = sotrsum.get(sotr, 0) + ws.cell(row=r, column=2).value
ws2.cell(row=1, column=1).value = 'Фамилия'
ws2.cell(row=1, column=2).value = 'Итог'
for k, v in sotrsum.items():
    ws2.append([k, v])
itogo = 0
for i in range(2, ws2.max_row + 1):
    itogo += int(ws2.cell(row=i, column=2).value)
ws2.append(['ИТОГО', str(itogo)])
print(sotrsum)
wb.save('Task_10-1.xlsx')

