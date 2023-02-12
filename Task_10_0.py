import openpyxl
from openpyxl import Workbook
#wb = Workbook()
#wb.save('Task_10_0.xlsx')
wb = openpyxl.load_workbook('Task_10_0.xlsx')
print(wb.sheetnames)
sk = []
for sh in wb.sheetnames:
    ws = wb[sh]
    n = ws.max_row * ws.max_column
    sk.append((sh, n))
wb.save('Task_10_0.xlsx')
print(sk)
sosk = sorted(sk, key=lambda x: (x[0], -x[1]))
print(sosk)




