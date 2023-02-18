import openpyxl
#from openpyxl import Workbook
#wb = Workbook()

wb = openpyxl.load_workbook('Task_10-2.xlsx')
ws1 = wb['Период_1']
ws2 = wb['Период_2']
ws3 = wb['Итоги']
fs1 = {}
fs2 = {}
for r in range(1, ws1.max_row + 1):
    f1 = ws1.cell(row=r, column=1).value
    fs1[f1] = fs1.get(f1, 0) + ws1.cell(row=r, column=2).value
for r in range(1, ws2.max_row + 1):
    f2 = ws2.cell(row=r, column=1).value
    fs2[f2] = fs1.get(f2, 0) + ws2.cell(row=r, column=2).value
#print(fs1, fs2)
fsi = {}
fsi.update(fs2)
#print(fsi)
for f in fs1.keys():
    fsi[f] = fsi.get(f, 0) + fs1[f]
sfsi = dict(sorted(fsi.items(), key=lambda x: x[0]))
#ws3.cell(row=1, column=1).value = 'Фамилия'
#ws3.cell(row=1, column=2).value = 'Сумм. выработка'
for k, v in sfsi.items():
    ws3.append([k, v])
itogo = 0
for i in range(1, ws3.max_row + 1):
    itogo += int(ws3.cell(row=i, column=2).value)
ws3.append(['ИТОГО', str(itogo)])
#print(fsi)
#print(sfsi)
wb.save('Task_10-2.xlsx')
