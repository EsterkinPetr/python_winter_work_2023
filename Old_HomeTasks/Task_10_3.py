import openpyxl
wb = openpyxl.load_workbook('Task_10-3.xlsx')
ws1 = wb['Лист_1']
ws2 = wb['Лист_2']
sprez = []
for row in range(2, ws1.max_row + 1):
    sprez.append(ws1.cell(row=row, column=2).value)
makzn = max(sprez)
mizn = min(sprez)
sar = sum(sprez) / len(sprez)
if len(sprez) % 2 == 0:
    med = (sprez[int(len(sprez)/2)] + sprez[int(len(sprez) / 2 - 1)]) / 2
else:
    med = sprez[int(len(sprez) // 2 - 1)]
ws2.append(['Минимальное значение', str(mizn)])
ws2.append(['Максимальное значение', str(makzn)])
ws2.append(['Среднее арифметическое', str(sar)])
ws2.append(['Медиана', str(med)])
#print(makzn, mizn, sar, med)
wb.save('Task_10-3.xlsx')

