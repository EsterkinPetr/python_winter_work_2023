import csv
from openpyxl import Workbook
wb = Workbook()
wb.save('Task_11_2.xlsx')
ws = wb.active
with open('Task_11-2.csv', 'r', encoding='utf-8') as ish:
    rows = list(csv.reader(ish))
    srows = sorted(rows, key=lambda x: (x[3], x[1], x[2]))
    print(srows)
    ws['A1'] = 'Номер'
    ws['B1'] = 'Фамилия'
    ws['C1'] = 'Имя'
    ws['D1'] = 'Организация'
    ws['E1'] = 'Зарплата'
    itogo = 0
    for i in range(len(srows)):
        for j in range(5):
            ws.cell(row=i + 2, column=j + 1).value = srows[i][j]
        itogo += int(srows[i][4])
    print(itogo)
    ws.append({1: 'Итого', 5: str(itogo)})
    wb.save('Task_11_2.xlsx')


